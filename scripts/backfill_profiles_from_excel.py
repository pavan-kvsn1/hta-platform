"""Backfill capability profiles for registry assets that have none, using the source spreadsheet.

16 of 204 assets carry parameter labels but zero capability_profiles, so they can
never appear in the pre-filtered "eligible masters" list (see
docs/scope/parameter-inventory-findings.md). The source spreadsheet
'MASTER LIST AS PER SOP & PARAMETER.xlsx' has a Range column that covers most of them.

This script emits an OVERLAY, not a mutation: docs/scope/registry-profile-backfill.json.
The registry stays the source of truth for what was actually extracted; the overlay
records what we can additionally justify from the spreadsheet, with provenance and a
confidence flag on every entry.

Important limitation: the spreadsheet has no accuracy or least-count column, so every
synthesized profile has an EMPTY buckets array. That restores range-based filtering for
these assets but not accuracy-based filtering. Profiles are marked so the UI can say so.

Usage:
    python scripts/backfill_profiles_from_excel.py
    python scripts/backfill_profiles_from_excel.py --check   # report only, no write
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

REF = Path("C:/Users/kcsva/OneDrive/Documents/HTACalibr8s/reference_docs/master_list")
DEFAULT_REGISTRY = REF / "master_details/master_instrment_registry.json"
DEFAULT_XLSX = REF / "MASTER LIST AS PER SOP & PARAMETER.xlsx"
DEFAULT_MAP = Path("docs/scope/parameter-normalization-map.json")
DEFAULT_OUT = Path("docs/scope/registry-profile-backfill.json")

SHEET = "MASTER LIST AS PER SOP"

# Range strings that carry no measurement range at all.
NON_RANGE = re.compile(r"^\s*(na|n/?a|-+|refer\b.*)\s*$", re.I)
# "400*400 mm" is an artifact dimension (a plate's size), not a measurement span.
DIMENSION = re.compile(r"^\s*[\d.]+\s*[*x×]\s*[\d.]+\s*\w*\s*$", re.I)

NUM = r"[-+]?\d[\d,]*\.?\d*"
RANGE_RE = re.compile(rf"^\s*({NUM})\s*([^\d\s]*[^\d]*?)\s*(?:to|-|–|~)\s*({NUM})\s*(.*?)\s*$", re.I)
SINGLE_RE = re.compile(rf"^\s*({NUM})\s*(.+?)\s*$")


def parse_number(text):
    """Parse a numeric literal, handling Indian digit grouping ('1,00,000' -> 100000)."""
    cleaned = text.replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_range(raw):
    """Return (parsed_dict_or_None, confidence, note).

    confidence: 'high'   - unambiguous 'A to B UNIT'
                'medium' - inferred (single-value range assumed to start at 0)
                'none'   - not a range; caller should skip
    """
    if raw is None:
        return None, "none", "no Range value in the spreadsheet"
    text = str(raw).strip()
    if not text:
        return None, "none", "empty Range cell"
    if NON_RANGE.match(text):
        if text.lower().startswith("refer"):
            return None, "none", f"Range points at an external document: {text!r}"
        return None, "none", f"Range is {text!r} - no numeric range declared"
    if DIMENSION.match(text):
        return None, "none", (
            f"Range {text!r} is an artifact dimension, not a measurement span"
        )

    # Scientific notation flattened by Excel: '103 to 10-3' means 10^3 to 10^-3.
    if re.match(r"^\s*10\s*(\d)\s*to\s*10\s*-\s*(\d)\s*$", text):
        m = re.match(r"^\s*10\s*(\d)\s*to\s*10\s*-\s*(\d)\s*$", text)
        hi = 10.0 ** int(m.group(1))
        lo = 10.0 ** (-int(m.group(2)))
        return (
            {"min": lo, "max": hi, "unit": None},
            "medium",
            f"{text!r} read as 10^{m.group(1)} to 10^-{m.group(2)}; superscripts were "
            f"lost in the spreadsheet. Unit is NOT stated - must be supplied manually.",
        )

    m = RANGE_RE.match(text)
    if m:
        lo = parse_number(m.group(1))
        hi = parse_number(m.group(3))
        unit = (m.group(4) or m.group(2) or "").strip() or None
        if lo is not None and hi is not None:
            if hi < lo:
                lo, hi = hi, lo
            return {"min": lo, "max": hi, "unit": unit}, "high", None

    m = SINGLE_RE.match(text)
    if m:
        value = parse_number(m.group(1))
        unit = m.group(2).strip() or None
        if value is not None:
            return (
                {"min": 0.0, "max": value, "unit": unit},
                "medium",
                f"{text!r} states only an upper bound; lower bound assumed to be 0",
            )

    return None, "none", f"could not parse Range {text!r}"


def load_excel(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[SHEET]
    rows = {}
    for row in sheet.iter_rows(values_only=True):
        asset = row[5]
        if not asset:
            continue
        rows[str(asset).strip()] = {
            "desc": row[1],
            "range": row[2],
            "parameter_raw": row[12],
        }
    return rows


def profile_count(asset):
    count = len(asset.get("capability_profiles") or [])
    for component in asset.get("components") or []:
        count += len(component.get("capability_profiles") or [])
        record = component.get("master_record") or {}
        count += len(record.get("capability_profiles") or [])
    return count


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--map", dest="map_path", type=Path, default=DEFAULT_MAP)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    registry = json.loads(args.registry.read_text(encoding="utf-8"))
    mapping = json.loads(args.map_path.read_text(encoding="utf-8"))
    excel = load_excel(args.xlsx)

    param_map = {
        k: v for k, v in mapping["parameters"].items() if not k.startswith("$")
    }
    alias_map = {
        k: v for k, v in mapping["assetLevelAliases"].items() if not k.startswith("$")
    }

    def resolve_parameter(name):
        """Resolve a label to standard names, via profile-level map then asset-level alias.

        Returns (standard_names, needs_new_standard).
        """
        entry = param_map.get(name)
        if entry:
            return [entry["standardName"]], False
        alias = alias_map.get(name)
        if alias is not None:
            targets = alias.get("resolvesTo") or []
            # An alias that resolves to nothing means the registry has no capability
            # coverage for this quantity at all - backfilling it requires creating a
            # new standard parameter, which is a product decision, not a data fix.
            return targets, not targets
        return [], True

    entries = []
    corrections = []

    for asset in registry["assets"]:
        if profile_count(asset) > 0:
            continue

        asset_no = asset.get("asset_no") or asset.get("asset_no_normalized")
        row = excel.get(str(asset_no).strip())
        params = asset.get("parameters") or []

        parsed, confidence, note = parse_range(row["range"] if row else None)

        # Flag places where the registry's own range_parsed disagrees with a clean
        # spreadsheet read - these are upstream parser bugs worth fixing at source.
        existing = asset.get("range_parsed")
        if parsed and existing:
            prior = existing[0]
            if (prior.get("max") != parsed["max"]) or (
                parsed["unit"] and prior.get("unit") != parsed["unit"]
            ):
                corrections.append(
                    {
                        "assetNo": asset_no,
                        "description": asset.get("instrument_desc"),
                        "rangeText": str(row["range"]).strip() if row else None,
                        "registryParsed": prior,
                        "spreadsheetRead": parsed,
                        "note": note
                        or "registry range_parsed disagrees with a clean read of the same text",
                    }
                )

        profiles = []
        needs_new_standard = []
        for raw_param in params:
            targets, new_standard = resolve_parameter(raw_param)
            if new_standard:
                needs_new_standard.append(raw_param)
            for standard in targets or [None]:
                profiles.append(
                    {
                        "sourceParameter": raw_param,
                        "standardName": standard,
                        "resolvable": standard is not None,
                        "role": "measuring",
                        "unit": parsed["unit"] if parsed else None,
                        "min": parsed["min"] if parsed else None,
                        "max": parsed["max"] if parsed else None,
                        "buckets": [],
                    }
                )

        # One Range cell cannot describe two different parameters. Applying the whole
        # span to each would overstate capability and turn a false-negative (master
        # invisible) into a false-positive (ineligible master shown as eligible), which
        # is strictly worse. These need a human to split the range.
        shared_range = parsed is not None and len(params) > 1

        # A range without a unit is not usable for filtering: comparing a UUC range in
        # mbar against a master range in unknown units is meaningless. Treat it as
        # blocked rather than ready, however confidently the numbers parsed.
        missing_unit = parsed is not None and not parsed.get("unit")

        if confidence == "none":
            status = "unresolvable"
        elif needs_new_standard:
            status = "needs-new-standard"
        elif missing_unit:
            status = "needs-unit"
        elif shared_range:
            status = "needs-range-split"
        else:
            status = "ready"

        entry = {
            "assetNo": asset_no,
            "description": asset.get("instrument_desc"),
            "excelRange": str(row["range"]).strip() if row and row["range"] else None,
            "excelParameter": row["parameter_raw"] if row else None,
            "rangeConfidence": confidence,
            "rangeNote": note,
            "status": status,
            "autoApplicable": status == "ready",
            "profiles": profiles,
        }
        if needs_new_standard:
            entry["needsNewStandardFor"] = needs_new_standard
        if shared_range:
            entry["rangeSplitWarning"] = (
                f"The single Range {entry['excelRange']!r} was applied to all "
                f"{len(params)} declared parameters. This overstates each of them and "
                f"must be split manually before use."
            )
        entries.append(entry)

    by_status = {}
    for e in entries:
        by_status.setdefault(e["status"], []).append(e)

    ready = by_status.get("ready", [])
    print(f"{len(entries)} assets with zero capability profiles")
    print(f"  ready to apply       : {len(ready)}")
    print(f"  needs range split    : {len(by_status.get('needs-range-split', []))}")
    print(f"  needs unit           : {len(by_status.get('needs-unit', []))}")
    print(f"  needs new standard   : {len(by_status.get('needs-new-standard', []))}")
    print(f"  unresolvable         : {len(by_status.get('unresolvable', []))}")
    print(f"  upstream parser bugs : {len(corrections)}")
    print()
    for e in by_status.get("needs-range-split", []):
        names = ", ".join(p["sourceParameter"] for p in e["profiles"])
        print(f"  RANGE SPLIT  {e['assetNo']:<16} {e['excelRange']!r} shared by: {names}")
    for e in by_status.get("needs-new-standard", []):
        print(f"  NEW STANDARD {e['assetNo']:<16} {', '.join(e['needsNewStandardFor'])}"
              f"  (range {e['excelRange']!r})")
    for e in by_status.get("needs-unit", []):
        print(f"  NEEDS UNIT   {e['assetNo']:<16} {e['excelRange']!r} parsed to "
              f"{e['profiles'][0]['min']}..{e['profiles'][0]['max']} but no unit is stated")
    for e in by_status.get("unresolvable", []):
        print(f"  UNRESOLVABLE {e['assetNo']:<16} {e['rangeNote']}")
    if corrections:
        print()
        for c in corrections:
            print(f"  PARSER BUG  {c['assetNo']:<16} {c['rangeText']!r}")
            print(f"              registry={c['registryParsed']}")
            print(f"              correct ={c['spreadsheetRead']}")

    if not args.check:
        payload = {
            "$comment": [
                "OVERLAY, not a registry replacement. Synthesized capability profiles for",
                "assets the extraction left with none, derived from the Range and PARAMETER",
                "columns of MASTER LIST AS PER SOP & PARAMETER.xlsx.",
                "",
                "Every profile here has an EMPTY buckets array: the spreadsheet carries no",
                "accuracy or least-count data. These assets can therefore participate in",
                "RANGE filtering but not ACCURACY filtering, and the UI must say so rather",
                "than implying full capability.",
                "",
                "'registryParserCorrections' lists places where the registry's own",
                "range_parsed is wrong versus a clean read of the same source text. Those",
                "should be fixed in the upstream extraction, not patched here.",
            ],
            "source": {
                "registry": str(args.registry.name),
                "spreadsheet": str(args.xlsx.name),
                "mapVersion": mapping.get("version"),
            },
            "zeroProfileAssetCount": len(entries),
            "readyCount": len(ready),
            "registryParserCorrections": corrections,
            "assets": entries,
        }
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        print(f"\nwritten to {args.out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
